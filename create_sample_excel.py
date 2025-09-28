#!/usr/bin/env python3

from openpyxl import Workbook
import os

def create_sample_decision_table():
    """Create a sample Drools decision table Excel file"""
    
    os.makedirs('rules', exist_ok=True)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "DecisionTable"
    
    headers = ["NAME", "CONDITION-1", "CONDITION-2", "ACTION-1", "ACTION-2"]
    
    templates = ["", "customer.age >= $1", "customer.category == \"$2\"", "discount.setPercentage($1)", "discount.setReason(\"$2\")"]
    
    data_rows = [
        ["Senior Discount", "65", "SENIOR", "15", "Senior citizen discount"],
        ["Student Discount", "18", "STUDENT", "10", "Student discount"],
        ["Premium Customer", "25", "PREMIUM", "20", "Premium customer discount"],
        ["Regular Customer", "18", "REGULAR", "5", "Regular customer discount"],
        ["VIP Customer", "21", "VIP", "25", "VIP customer discount"]
    ]
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    for col, template in enumerate(templates, 1):
        ws.cell(row=2, column=col, value=template)
    
    for row_idx, data_row in enumerate(data_rows, 3):
        for col, value in enumerate(data_row, 1):
            ws.cell(row=row_idx, column=col, value=value)
    
    filename = "rules/customer_discount_rules.xlsx"
    wb.save(filename)
    print(f"Created sample decision table: {filename}")
    
    return filename

def create_product_rules_table():
    """Create another sample decision table for product rules"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "ProductRules"
    
    headers = ["NAME", "CONDITION-1", "CONDITION-2", "ACTION-1"]
    
    templates = ["", "product.category == \"$1\"", "order.quantity >= $2", "shipping.setFree($1)"]
    
    data_rows = [
        ["Electronics Free Shipping", "ELECTRONICS", "100", "true"],
        ["Books Free Shipping", "BOOKS", "50", "true"],
        ["Clothing Bulk Order", "CLOTHING", "200", "true"],
        ["Home & Garden", "HOME", "150", "true"]
    ]
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    for col, template in enumerate(templates, 1):
        ws.cell(row=2, column=col, value=template)
    
    for row_idx, data_row in enumerate(data_rows, 3):
        for col, value in enumerate(data_row, 1):
            ws.cell(row=row_idx, column=col, value=value)
    
    filename = "rules/product_shipping_rules.xlsx"
    wb.save(filename)
    print(f"Created sample decision table: {filename}")
    
    return filename

if __name__ == "__main__":
    print("Creating sample Drools decision tables...")
    
    file1 = create_sample_decision_table()
    file2 = create_product_rules_table()
    
    print(f"\nSample files created:")
    print(f"1. {file1}")
    print(f"2. {file2}")
    print("\nThese files contain sample Drools decision tables that can be used to test the Rules Manager application.")
